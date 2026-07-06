"""
excel_db.py  –  Database Layer
รองรับ 2 โหมด:
  - LOCAL: อ่าน/เขียน Excel (roon_database.xlsx)
  - CLOUD: อ่าน/เขียน Google Sheets
ตรวจจาก environment variable USE_GSHEETS=true หรือ st.secrets มี gcp_service_account
"""
import os
import pandas as pd

# ── ตรวจว่าควรใช้ Google Sheets หรือไม่ ───────────────────────
def _use_gsheets() -> bool:
    """ถ้ามี gcp_service_account ใน secrets → ใช้ Google Sheets เสมอ"""
    if os.environ.get("USE_GSHEETS", "").lower() == "true":
        return True
    try:
        import streamlit as st
        if "gcp_service_account" in st.secrets:
            return True
    except Exception:
        pass
    return False


# ── PROXY: เลือก backend อัตโนมัติ ───────────────────────────
def _backend():
    if _use_gsheets():
        import gsheets_db as gs
        return gs
    else:
        import sys
        return sys.modules[__name__]._local_backend()


class _LocalBackend:
    """Excel backend (เดิม)"""

    def __init__(self):
        from config import DB_PATH
        self.DB_PATH = DB_PATH

    def init_workbook(self):
        from config import ALL_SHEETS
        from openpyxl import load_workbook, Workbook
        if os.path.exists(self.DB_PATH):
            # ตรวจก่อนว่าไฟล์ไม่เสียหาย
            try:
                wb = load_workbook(self.DB_PATH)
                changed = False
                for sheet in ALL_SHEETS:
                    if sheet not in wb.sheetnames:
                        wb.create_sheet(sheet)
                        changed = True
                if changed:
                    wb.save(self.DB_PATH)
                wb.close()
            except Exception:
                # ไฟล์เสียหาย — สร้างใหม่
                import os as _os
                try: _os.remove(self.DB_PATH)
                except: pass
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                for sheet in ALL_SHEETS:
                    wb.create_sheet(sheet)
                wb.save(self.DB_PATH)
                wb.close()
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            for sheet in ALL_SHEETS:
                wb.create_sheet(sheet)
            wb.save(self.DB_PATH)
            wb.close()

    def read_sheet(self, sheet_name: str) -> pd.DataFrame:
        self.init_workbook()
        try:
            df = pd.read_excel(self.DB_PATH, sheet_name=sheet_name, dtype=str)
            return df.fillna("")
        except Exception:
            return pd.DataFrame()

    def write_sheet(self, sheet_name: str, df: pd.DataFrame):
        self.init_workbook()
        from openpyxl import load_workbook
        wb = load_workbook(self.DB_PATH)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        wb.save(self.DB_PATH)
        wb.close()

    def append_row(self, sheet_name: str, row_dict: dict):
        df = self.read_sheet(sheet_name)
        new_row = pd.DataFrame([row_dict])
        df = pd.concat([df, new_row], ignore_index=True)
        self.write_sheet(sheet_name, df)

    def update_row(self, sheet_name: str, id_col: str, id_value: str, updated_dict: dict):
        df = self.read_sheet(sheet_name)
        mask = df[id_col].astype(str) == str(id_value)
        for key, val in updated_dict.items():
            df.loc[mask, key] = str(val) if val is not None else ""
        self.write_sheet(sheet_name, df)

    def delete_row(self, sheet_name: str, id_col: str, id_value: str):
        df = self.read_sheet(sheet_name)
        df = df[df[id_col].astype(str) != str(id_value)]
        self.write_sheet(sheet_name, df)


_local = _LocalBackend()

def _local_backend():
    return _local


# ── PUBLIC API (เรียกใช้จากทุก module) ────────────────────────
def init_workbook():
    if _use_gsheets():
        try:
            import gsheets_db as gs
            gs.init_workbook()
        except Exception as e:
            import streamlit as st
            st.warning(f"⚠️ init Google Sheets: {e}")
    else:
        # Local mode — ถ้าไฟล์เสียหายให้ข้ามไป
        try:
            _local.init_workbook()
        except Exception:
            pass

def read_sheet(sheet_name: str) -> pd.DataFrame:
    if _use_gsheets():
        import gsheets_db as gs
        return gs.read_sheet(sheet_name)
    return _local.read_sheet(sheet_name)

def write_sheet(sheet_name: str, df: pd.DataFrame):
    if _use_gsheets():
        import gsheets_db as gs
        gs.write_sheet(sheet_name, df)
    else:
        _local.write_sheet(sheet_name, df)

def append_row(sheet_name: str, row_dict: dict):
    if _use_gsheets():
        import gsheets_db as gs
        gs.append_row(sheet_name, row_dict)
    else:
        _local.append_row(sheet_name, row_dict)

def update_row(sheet_name: str, id_col: str, id_value: str, updated_dict: dict):
    if _use_gsheets():
        import gsheets_db as gs
        gs.update_row(sheet_name, id_col, id_value, updated_dict)
    else:
        _local.update_row(sheet_name, id_col, id_value, updated_dict)

def delete_row(sheet_name: str, id_col: str, id_value: str):
    if _use_gsheets():
        import gsheets_db as gs
        gs.delete_row(sheet_name, id_col, id_value)
    else:
        _local.delete_row(sheet_name, id_col, id_value)
